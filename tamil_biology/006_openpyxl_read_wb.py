from openpyxl import load_workbook
import collections as col

wb = load_workbook(filename = 'RangeBook.xlsx')

print(wb.sheetnames)

sheet_ranges = wb['One']

#print(sheet_ranges['A2'].value)

QuizRecord = col.namedtuple('QuizRecord', 'name, price, quantity')

QuizRecordsDb = []

for row in range(2, 5, 1):
    print(sheet_ranges['A' + str(row)].value, sheet_ranges['B' + str(row)].value, sheet_ranges['C' + str(row)].value, sep=',')
    # print(sheet_ranges['B' + str(row)].value)
    # print(sheet_ranges['C' + str(row)].value)
    quizRecord = QuizRecord(sheet_ranges['A' + str(row)].value, sheet_ranges['B' + str(row)].value, sheet_ranges['C' + str(row)].value)
    QuizRecordsDb.append(quizRecord)

# print(sheet_ranges['A' + str(1)].value)
# print(sheet_ranges['B' + str(1)].value)
# print(sheet_ranges['C' + str(1)].value)

for x in QuizRecordsDb:
  print(x)
